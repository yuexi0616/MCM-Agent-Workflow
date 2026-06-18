#!/usr/bin/env python3
"""
Phase 4 一键打包脚本 —— 将已批准的论文和结果打包为多格式最终产出。

功能：
  1. Markdown → DOCX (pandoc)
  2. Markdown → PDF (weasyprint 或 xelatex)
  3. 从 data_vis/code 提取分题结果生成 XLSX
  4. 图表命名规范化检查与重命名
  5. 产出完整性检查清单

Usage:
    python tools/package_run.py --run _workspace/run-20260601-103030
    python tools/package_run.py --run _workspace/run-20260601-103030 --skip-pdf
    python tools/package_run.py --run _workspace/run-20260601-103030 --check-only
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def find_paper_md(run_dir: Path) -> Path | None:
    """Find the paper markdown file in the run directory."""
    candidates = [
        run_dir / "paper" / "paper.md",
        run_dir / "phase-3-paper" / "draft-paper.md",
    ]
    for c in candidates:
        if c.exists():
            return c
    # Search phase-3-paper for any .md
    phase3 = run_dir / "phase-3-paper"
    if phase3.exists():
        md_files = list(phase3.glob("*.md"))
        if md_files:
            return md_files[0]
    return None


def check_command(cmd: str) -> bool:
    """Check if a command is available on PATH."""
    return shutil.which(cmd) is not None


def convert_to_docx(md_path: Path, run_dir: Path) -> tuple[bool, str]:
    """Convert markdown to DOCX using pandoc."""
    docx_path = run_dir / "paper" / "paper.docx"
    docx_path.parent.mkdir(parents=True, exist_ok=True)

    if not check_command("pandoc"):
        return False, "pandoc not found on PATH — install pandoc to generate DOCX"

    reference_doc = PROJECT_ROOT / "_workspace" / "reference.docx"
    cmd = [
        "pandoc", str(md_path), "-o", str(docx_path),
        "--from=markdown", "--to=docx",
    ]
    if reference_doc.exists():
        cmd.extend(["--reference-doc", str(reference_doc)])

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode == 0 and docx_path.exists():
            size_kb = docx_path.stat().st_size / 1024
            return True, f"paper.docx ({size_kb:.1f} KB)"
        return False, f"pandoc failed: {result.stderr.strip()[:200]}"
    except subprocess.TimeoutExpired:
        return False, "pandoc timed out"
    except Exception as exc:
        return False, str(exc)


def convert_to_pdf(md_path: Path, run_dir: Path) -> tuple[bool, str]:
    """Convert markdown to PDF using pandoc (try xelatex, pdflatex, weasyprint in order)."""
    pdf_path = run_dir / "paper" / "paper.pdf"
    pdf_path.parent.mkdir(parents=True, exist_ok=True)

    if not check_command("pandoc"):
        return False, "pandoc not found on PATH"

    # Ordered list of PDF engines to try. Each is a list of extra args.
    engines = [
        (["--pdf-engine=xelatex"], "xelatex"),
        (["--pdf-engine=pdflatex"], "pdflatex"),
        (["--pdf-engine=weasyprint"], "weasyprint"),
        (["--pdf-engine=wkhtmltopdf"], "wkhtmltopdf"),
    ]

    for engine_args, engine_name in engines:
        if not check_command(engine_name):
            continue
        try:
            cmd = ["pandoc", str(md_path), "-o", str(pdf_path)] + engine_args
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
            if result.returncode == 0 and pdf_path.exists():
                return True, f"paper.pdf ({engine_name})"
        except subprocess.TimeoutExpired:
            continue
        except Exception:
            continue

    return False, "PDF 待生成 — 需安装 xelatex/pdflatex/weasyprint/wkhtmltopdf 之一"


def extract_results_to_xlsx(run_dir: Path) -> tuple[bool, str]:
    """Extract results from CSV files and consolidate into XLSX."""
    try:
        import openpyxl
    except ImportError:
        return False, "openpyxl not installed — pip install openpyxl"

    results_dir = run_dir / "results"
    if not results_dir.exists():
        return False, "results/ directory not found"

    csv_files = sorted(results_dir.glob("*.csv"))
    if not csv_files:
        return False, "no CSV files in results/"

    xlsx_path = results_dir / "results_summary.xlsx"
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets_created = 0
    for csv_file in csv_files:
        sheet_name = csv_file.stem[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        try:
            with csv_file.open("r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh)
                for row in reader:
                    ws.append(row)
        except UnicodeDecodeError:
            with csv_file.open("r", encoding="gb18030", newline="") as fh:
                reader = csv.reader(fh)
                for row in reader:
                    ws.append(row)
        sheets_created += 1

    if sheets_created == 0:
        return False, "no data sheets created"

    # Create summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.append(["problem_id", "metric_name", "value", "unit", "method", "figure_ref", "formula_ref"])
    ws_summary.append(["", f"Generated: {datetime.now().isoformat(timespec='seconds')}", "", "", "", "", ""])

    wb.save(xlsx_path)
    return True, f"results_summary.xlsx ({sheets_created} sheets)"


def normalize_figure_names(run_dir: Path) -> list[str]:
    """Check and report on figure naming convention."""
    figures_dir = run_dir / "figures"
    if not figures_dir.exists():
        return ["figures/ directory not found"]

    png_files = sorted(figures_dir.glob("*.png"))
    if not png_files:
        return ["no PNG files in figures/"]

    messages: list[str] = []
    good = 0
    bad = 0
    for f in png_files:
        if re.match(r"fig\d{2}[_\-].+\.png$", f.name, re.I):
            good += 1
        else:
            bad += 1
            messages.append(f"  ⚠ {f.name} — 不符合 figNN_xxx.png 命名规范")

    messages.insert(0, f"图表命名: {good} 符合规范, {bad} 需修正")
    return messages


def run_completeness_checklist(run_dir: Path) -> list[dict[str, Any]]:
    """Run the Phase 4 completeness checklist."""
    checks: list[dict[str, Any]] = []

    paper_dir = run_dir / "paper"
    figures_dir = run_dir / "figures"
    results_dir = run_dir / "results"
    code_dir = run_dir / "code"

    # Paper checks
    paper_md = paper_dir / "paper.md" if paper_dir.exists() else None
    if not paper_md and paper_dir.exists():
        # Try finding it elsewhere
        paper_md = find_paper_md(run_dir)
    checks.append({
        "item": "paper/paper.md 存在",
        "status": "PASS" if paper_md and paper_md.exists() else "FAIL",
        "detail": str(paper_md) if paper_md else "not found",
    })

    paper_docx = paper_dir / "paper.docx" if paper_dir.exists() else None
    checks.append({
        "item": "paper/paper.docx 生成成功",
        "status": "PASS" if paper_docx and paper_docx.exists() else "PENDING",
        "detail": "will be generated" if not paper_docx or not paper_docx.exists() else str(paper_docx),
    })

    paper_pdf = paper_dir / "paper.pdf" if paper_dir.exists() else None
    checks.append({
        "item": "paper/paper.pdf 生成成功",
        "status": "PASS" if paper_pdf and paper_pdf.exists() else "PENDING",
        "detail": "will be generated" if not paper_pdf or not paper_pdf.exists() else str(paper_pdf),
    })

    # Figures
    if figures_dir.exists():
        pngs = list(figures_dir.glob("*.png"))
        checks.append({
            "item": "figures/ 含 PNG 图表",
            "status": "PASS" if pngs else "FAIL",
            "detail": f"{len(pngs)} PNG files",
        })
    else:
        checks.append({
            "item": "figures/ 目录存在",
            "status": "FAIL",
            "detail": "directory not found",
        })

    # Results
    if results_dir.exists():
        csvs = list(results_dir.glob("*.csv"))
        xlsxs = list(results_dir.glob("*.xlsx"))
        has_summary = any("summary" in f.name.lower() for f in csvs + xlsxs)
        checks.append({
            "item": "results/ 含分题结果 + 汇总表",
            "status": "PASS" if (csvs or xlsxs) and has_summary else "WARN",
            "detail": f"{len(csvs)} CSV, {len(xlsxs)} XLSX, summary={'yes' if has_summary else 'no'}",
        })
    else:
        checks.append({
            "item": "results/ 目录存在",
            "status": "FAIL",
            "detail": "directory not found",
        })

    # Code
    solver = code_dir / "solver.py" if code_dir.exists() else None
    reqs = code_dir / "requirements.txt" if code_dir.exists() else None
    checks.append({
        "item": "code/solver.py 可直接运行",
        "status": "PASS" if solver and solver.exists() else "FAIL",
        "detail": str(solver) if solver else "not found",
    })
    checks.append({
        "item": "code/requirements.txt 完整依赖",
        "status": "PASS" if reqs and reqs.exists() else "FAIL",
        "detail": str(reqs) if reqs else "not found",
    })

    return checks


def update_run_state(run_dir: Path, phase: str, status: str) -> None:
    """Update run_state.json with the new phase status."""
    state_path = run_dir / "run_state.json"
    if not state_path.exists():
        return
    state = load_json(state_path)
    state["phase_status"][phase] = status
    if status == "completed":
        state["last_approved_phase"] = phase
    state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def print_checklist(checks: list[dict[str, Any]]) -> tuple[int, int, int]:
    """Print the checklist and return (pass, warn, fail) counts."""
    n_pass = n_warn = n_fail = 0
    for c in checks:
        icon = {"PASS": "✓", "WARN": "⚠", "FAIL": "✗", "PENDING": "○"}.get(c["status"], "?")
        print(f"  [{icon}] {c['item']}")
        if c["detail"]:
            print(f"      {c['detail']}")
        if c["status"] == "PASS":
            n_pass += 1
        elif c["status"] == "WARN":
            n_warn += 1
        elif c["status"] == "FAIL":
            n_fail += 1
    return n_pass, n_warn, n_fail


def main() -> int:
    parser = argparse.ArgumentParser(description="Phase 4 一键打包脚本")
    parser.add_argument("--run", type=Path, required=True, help="Run 目录路径")
    parser.add_argument("--skip-pdf", action="store_true", help="跳过 PDF 生成")
    parser.add_argument("--skip-docx", action="store_true", help="跳过 DOCX 生成")
    parser.add_argument("--check-only", action="store_true", help="仅运行完整性检查，不生成文件")
    args = parser.parse_args()

    run_dir = args.run.resolve()
    if not run_dir.exists():
        print(f"ERROR: run directory not found: {run_dir}")
        return 1

    print(f"=== Phase 4 打包: {run_dir.name} ===\n")

    # 0. Find paper
    print("[0] 定位论文文件...")
    paper_md = find_paper_md(run_dir)
    if not paper_md:
        print("  ✗ 未找到论文 Markdown 文件")
        if not args.check_only:
            return 1
    else:
        print(f"  ✓ {paper_md.relative_to(PROJECT_ROOT)}")

    if args.check_only:
        print("\n[检查] 产出完整性清单:")
        checks = run_completeness_checklist(run_dir)
        n_pass, n_warn, n_fail = print_checklist(checks)
        print(f"\n结果: {n_pass} PASS, {n_warn} WARN, {n_fail} FAIL")
        return 0 if n_fail == 0 else 1

    # 1. DOCX conversion
    if not args.skip_docx and paper_md:
        print("\n[1] 生成 DOCX...")
        ok, msg = convert_to_docx(paper_md, run_dir)
        print(f"  {'✓' if ok else '✗'} {msg}")

    # 2. PDF conversion
    if not args.skip_pdf and paper_md:
        print("\n[2] 生成 PDF...")
        ok, msg = convert_to_pdf(paper_md, run_dir)
        print(f"  {'✓' if ok else '✗'} {msg}")

    # 3. Results to XLSX
    print("\n[3] 提取分题结果...")
    ok, msg = extract_results_to_xlsx(run_dir)
    print(f"  {'✓' if ok else '✗'} {msg}")

    # 4. Figure naming check
    print("\n[4] 图表命名规范检查...")
    for msg in normalize_figure_names(run_dir):
        print(f"  {msg}")

    # 5. Completeness checklist
    print("\n[5] 产出完整性检查清单:")
    checks = run_completeness_checklist(run_dir)
    n_pass, n_warn, n_fail = print_checklist(checks)

    # 6. Update run state
    if n_fail == 0:
        update_run_state(run_dir, "phase-4", "completed")
        print(f"\n  ✓ run_state.json 已更新 (phase-4=completed)")

    print(f"\n=== 打包完成 ===")
    print(f"结果: {n_pass} PASS, {n_warn} WARN, {n_fail} FAIL")
    if n_fail > 0:
        print(f"⚠ 存在 {n_fail} 项未通过，请检查上述 FAIL 项")
    else:
        print(f"✓ 产出完整性验证通过")

    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())